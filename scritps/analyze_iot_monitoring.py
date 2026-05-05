import os, sys, argparse, warnings
import pandas as pd
import numpy as np
from pathlib import Path

warnings.filterwarnings("ignore")

# Konfiguration

NOISE = ["SURICATA Ethertype unknown", "SURICATA STREAM", "SURICATA TCP"]

SCENARIOS_NORMAL = {
    "S1_baseline":           ("Baseline", "–"),
    "S2_suricata_no_attack": ("Suricata", "Suricata"),
    "S3_zeek_no_attack":     ("Zeek",     "Zeek"),
    "S4_hybrid_no_attack":   ("Hybrid",   "Suricata+Zeek"),
}
SCENARIOS_ATTACK = {
    "S5_suricata_attack": ("Suricata", "Suricata"),
    "S6_zeek_attack":     ("Zeek",     "Zeek"),
    "S7_hybrid_attack":   ("Hybrid",   "Suricata+Zeek"),
}
ATTACKS = ["portscan", "dos", "mqtt"]

# Hilfsfunktionen

def find_runs(path):
    return sorted([p for p in Path(path).iterdir() if p.is_dir() and "run" in p.name]) if Path(path).exists() else []

def load_csv(path):
    try:
        return pd.read_csv(path, parse_dates=["timestamp"]) if Path(path).exists() else None
    except:
        return None

def load_timeline(path):
    try:
        df = pd.read_csv(path, parse_dates=["timestamp"])
        return {r["event"]: r["timestamp"] for _, r in df.iterrows()}
    except:
        return {}

def agg(dfs, col):
    if not dfs:
        return {"mean": np.nan, "max": np.nan}
    vals = pd.concat([d[col] for d in dfs if col in d.columns])
    return {"mean": round(vals.mean(), 3), "max": round(vals.max(), 3)}

def during_attack(df, tl):
    if df is None or df.empty or "ATTACK_START" not in tl:
        return df
    return df[(df["timestamp"] >= tl["ATTACK_START"]) & (df["timestamp"] <= tl.get("ATTACK_END", df["timestamp"].max()))]

def load_alerts(path):
    try:
        df = pd.read_csv(path, parse_dates=["timestamp"])
        noise = df["signature"].str.contains("|".join(NOISE), na=False, case=False)
        return df[~noise], df[noise]
    except:
        return pd.DataFrame(), pd.DataFrame()

# Ressourcen Normalbetrieb

def analyze_normal(gw_root, mon_root):
    rows = []
    for sd, (label, monitoring) in SCENARIOS_NORMAL.items():
        gw_dfs  = [load_csv(r/"gateway_metrics.csv")    for r in find_runs(gw_root/sd)  if load_csv(r/"gateway_metrics.csv")    is not None]
        mon_dfs = [load_csv(r/"monitoring_metrics.csv") for r in find_runs(mon_root/sd) if load_csv(r/"monitoring_metrics.csv") is not None]
        rows.append({
            "Szenario": label, "Monitoring": monitoring,
            "GW CPU Ø (%)":     agg(gw_dfs,  "cpu_percent")["mean"],
            "GW CPU Max (%)":   agg(gw_dfs,  "cpu_percent")["max"],
            "GW RAM Ø (MB)":    agg(gw_dfs,  "ram_used_mb")["mean"],
            "MON CPU Ø (%)":    agg(mon_dfs, "cpu_percent")["mean"],
            "MON CPU Max (%)":  agg(mon_dfs, "cpu_percent")["max"],
            "MON RAM Ø (MB)":   agg(mon_dfs, "ram_used_mb")["mean"],
            "MON RAM Max (MB)": agg(mon_dfs, "ram_used_mb")["max"],
            "Net Ø (KB)":       round(pd.concat([d["bytes_recv_diff"] for d in mon_dfs if "bytes_recv_diff" in d]).sum() / 1024, 1) if mon_dfs else np.nan,
        })
    return pd.DataFrame(rows)

# Ressourcen unter Angriff

def analyze_attack_resources(gw_root, mon_root):
    rows = []
    for sd, (label, _) in SCENARIOS_ATTACK.items():
        snum = list(SCENARIOS_ATTACK.keys()).index(sd) + 5
        for attack in ATTACKS:
            gw_dfs, mon_dfs = [], []
            for r in find_runs(gw_root/sd/attack):
                tl = load_timeline(r/"timeline.csv")
                df = load_csv(r/"gateway_metrics.csv")
                if df is not None and tl:
                    att = during_attack(df, tl)
                    if not att.empty:
                        gw_dfs.append(att)
            for r in find_runs(mon_root/sd/attack):
                tl = load_timeline((gw_root/sd/attack/r.name/"timeline.csv"))
                df = load_csv(r/"monitoring_metrics.csv")
                if df is not None:
                    att = during_attack(df, tl) if tl else df
                    if not att.empty:
                        mon_dfs.append(att)
            rows.append({
                "Szenario": f"S{snum} – {label}", "Angriffstyp": attack.upper(),
                "GW CPU Ø (%)":     agg(gw_dfs,  "cpu_percent")["mean"],
                "GW CPU Max (%)":   agg(gw_dfs,  "cpu_percent")["max"],
                "GW RAM Ø (MB)":    agg(gw_dfs,  "ram_used_mb")["mean"],
                "MON CPU Ø (%)":    agg(mon_dfs, "cpu_percent")["mean"],
                "MON CPU Max (%)":  agg(mon_dfs, "cpu_percent")["max"],
                "MON RAM Ø (MB)":   agg(mon_dfs, "ram_used_mb")["mean"],
                "MON RAM Max (MB)": agg(mon_dfs, "ram_used_mb")["max"],
            })
    return pd.DataFrame(rows)

# Erkennungsmetriken

def analyze_detection(gw_root, mon_root):
    rows = []
    for sd, (label, monitoring) in SCENARIOS_ATTACK.items():
        snum = list(SCENARIOS_ATTACK.keys()).index(sd) + 5
        for attack in ATTACKS:
            tp, fp, fn, rt = [], [], [], []
            has_sur  = "suricata" in monitoring.lower()
            has_zeek = "zeek"     in monitoring.lower()

            for r in find_runs(mon_root/sd/attack):
                tl = load_timeline(gw_root/sd/attack/r.name/"timeline.csv")

                if has_sur:
                    real, _ = load_alerts(r/"suricata_alerts.csv")
                    if tl and "ATTACK_START" in tl:
                        win = real[(real["timestamp"] >= tl["ATTACK_START"]) &
                                   (real["timestamp"] <= tl.get("ATTACK_END", real["timestamp"].max()))] if not real.empty else pd.DataFrame()
                        detected = not win.empty
                        tp.append(1 if detected else 0)
                        fn.append(0 if detected else 1)
                        fp.append(len(real[real["timestamp"] < tl["ATTACK_START"]]) if not real.empty else 0)
                        if detected:
                            rt.append((win["timestamp"].min() - tl["ATTACK_START"]).total_seconds())

                if has_zeek:
                    zeek_dir = r/"zeek_logs"
                    conn     = zeek_dir/"conn.log"
                    mqtt_log = zeek_dir/"mqtt_publish.log"
                    if conn.exists():
                        size     = conn.stat().st_size / 1e6
                        detected = (size > 5 if attack == "portscan" else
                                    (mqtt_log.exists() and mqtt_log.stat().st_size > 100_000) if attack == "dos" else
                                    (mqtt_log.exists() and mqtt_log.stat().st_size > 50_000))
                        if has_sur:
                            if tp and tp[-1] == 0 and detected:
                                tp[-1], fn[-1] = 1, 0
                        else:
                            tp.append(1 if detected else 0)
                            fn.append(0 if detected else 1)
                            fp.append(0)

            tp_s, fn_s, fp_s = sum(tp), sum(fn), sum(fp)
            recall    = tp_s / (tp_s + fn_s) if tp_s + fn_s > 0 else 0
            precision = tp_s / (tp_s + fp_s) if tp_s + fp_s > 0 else 0
            f1        = 2 * precision * recall / (precision + recall) if precision + recall > 0 else 0

            rows.append({
                "Szenario": f"S{snum} – {label}", "Monitoring": monitoring,
                "Angriffstyp": attack.upper(),
                "TP": tp_s, "FP": fp_s, "FN": fn_s,
                "Recall": round(recall, 3), "Precision": round(precision, 3),
                "F1-Score": round(f1, 3),
                "FP-Rate": round(fp_s / max(tp_s + fp_s, 1), 3),
                "Reaktionszeit Ø (s)": round(np.mean(rt), 2) if rt else "n/a",
            })
    return pd.DataFrame(rows)

# MQTT Frequenz

def analyze_mqtt(gw_root):
    rows = []
    for sd, (label, _) in {**SCENARIOS_NORMAL, **SCENARIOS_ATTACK}.items():
        is_attack = sd in SCENARIOS_ATTACK
        paths     = [gw_root/sd/a for a in ATTACKS] if is_attack else [gw_root/sd]
        for path in paths:
            attack = path.name if is_attack else "–"
            normal, attack_freq = [], []
            for r in find_runs(path):
                csv = r/"mqtt_messages.csv"
                if not csv.exists():
                    continue
                try:
                    df = pd.read_csv(csv)
                    tl = load_timeline(r/"timeline.csv")
                    if tl and "ATTACK_START" in tl and "timestamp" in df.columns:
                        df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
                        pre     = df[df["timestamp"] < tl["ATTACK_START"]]
                        dur     = df[(df["timestamp"] >= tl["ATTACK_START"]) & (df["timestamp"] <= tl.get("ATTACK_END", df["timestamp"].max()))]
                        pre_dur = max((tl["ATTACK_START"] - tl.get("RUN_START", tl["ATTACK_START"])).total_seconds(), 1)
                        att_dur = max((tl.get("ATTACK_END", tl["ATTACK_START"]) - tl["ATTACK_START"]).total_seconds(), 1)
                        normal.append(len(pre) / pre_dur)
                        attack_freq.append(len(dur) / att_dur)
                    else:
                        normal.append(len(df) / 60)
                except:
                    continue
            rows.append({
                "Szenario": label, "Angriffstyp": attack.upper(),
                "Nachrichten/s (Normal)":  round(np.mean(normal),       2) if normal       else np.nan,
                "Nachrichten/s (Angriff)": round(np.mean(attack_freq),  2) if attack_freq  else np.nan,
            })
    return pd.DataFrame(rows)

# Excel Export

def write_excel(path, sheets):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            if df is not None and not df.empty:
                df.to_excel(writer, sheet_name=name[:31], index=False)
                ws = writer.sheets[name[:31]]
                from openpyxl.styles import PatternFill, Font, Alignment
                fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                for cell in ws[1]:
                    cell.fill      = fill
                    cell.font      = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center")
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col) + 4, 30)

# Hauptprogramm

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--results", "-r", default=".")
    args = parser.parse_args()

    root     = Path(args.results)
    gw_root  = root/"results_gateway_pi"
    mon_root = root/"results_monitoring_pi"

    if not gw_root.exists():
        print(f"Fehler: {gw_root} nicht gefunden")
        sys.exit(1)

    print("Auswertung läuft...")

    df_normal = analyze_normal(gw_root, mon_root)
    df_attack = analyze_attack_resources(gw_root, mon_root)
    df_detect = analyze_detection(gw_root, mon_root)
    df_mqtt   = analyze_mqtt(gw_root)

    write_excel(root/"evaluation_summary.xlsx", {
        "T1 Ressourcen Normal":  df_normal,
        "T2 Ressourcen Angriff": df_attack,
        "T3 Erkennungsmetriken": df_detect,
        "T4 MQTT-Frequenz":      df_mqtt,
    })

    print(f"Fertig! Excel: {root/'evaluation_summary.xlsx'}")
    print("Grafiken erstellen: python generate_plots.py --excel evaluation_summary.xlsx")

if __name__ == "__main__":
    main()
